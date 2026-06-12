"""
cca_merge.py — CCA Script
Merges an intake template (e.g. "Time Tested.xlsx") into Gallo_Metadata_Extract.xlsx.

Maps intake template column names to Gallo column names, then appends rows.
New columns in the intake that have no Gallo equivalent are ignored.
Gallo columns with no intake equivalent are left blank for appended rows.

Usage:
    python3 cca_merge.py --intake "Time Tested.xlsx" --gallo "Gallo_Metadata_Extract.xlsx"

Optional flags:
    --intake-sheet   Sheet name in intake file (default: first sheet)
    --gallo-sheet    Sheet name in Gallo file   (default: "Metadata")
    --output         Output file path           (default: overwrites Gallo file)
    --dry-run        Print mapped rows without writing
"""

import argparse
import openpyxl
from datetime import datetime, time as dt_time
import sys
import os

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN MAPPING: intake template header → Gallo header
# Add new mappings here as new intake templates are encountered.
# Keys are lowercased + stripped for fuzzy matching.
# ─────────────────────────────────────────────────────────────────────────────
COLUMN_MAP = {
    # Direct / near-direct matches
    "album title":                              "Album title",
    "album artist":                             "Album artist",
    "label":                                    "Label",
    "original release date":                    "Original Release Date",
    "release date":                             "Release Date",
    "track name":                               "Track name",

    # Asterisk variants (ISRC*, Language*, etc.)
    "isrc*":                                    "ISRC",
    "isrc":                                     "ISRC",
    "language*":                                "Language",
    "language":                                 "Language",

    # Track fields
    "track number":                             "#",
    "track artist (if different from album artist)": "Track artist",
    "track artist":                             "Track artist",
    "track duration":                           "Duration",
    "duration":                                 "Duration",

    # Featured artist → Performing contributors
    "featuredartist (if any)":                  "Performing contributors",
    "featured artist (if any)":                 "Performing contributors",
    "featured artist":                          "Performing contributors",

    # Genre — mapped to both album and track genre
    # (handled specially in code below — see DUAL_MAP)
    "genre":                                    "Album genre",

    # Credits
    "composer(s)":                              "Writers / Composers",
    "composers":                                "Writers / Composers",
    "writer(s)":                                "Writers / Composers",
    "writers":                                  "Writers / Composers",
    "producer(s)":                              "Producers",
    "producers":                                "Producers",
    "publisher(s)":                             "Publishers",
    "publishers":                               "Publishers",

    # Identifiers
    "upc":                                      "Barcode",
    "barcode":                                  "Barcode",
    "reference catalogue number":               "Cat #",
    "cat #":                                    "Cat #",
    "catalogue number":                         "Cat #",
    "external / client identifier":             "Album ID",
    "external/client identifier":               "Album ID",

    # Explicit / lyrics rating
    "lyrical content rating":                   "Explicit",
    "explicit":                                 "Explicit",
    "explicit lyrics":                          "Explicit lyrics",

    # P/C lines
    "℗ p line":                                 "Album ℗ line",
    "p line":                                   "Album ℗ line",
    "℗ line":                                   "Album ℗ line",
    "© c line":                                 "Album © line",
    "c line":                                   "Album © line",
    "© line":                                   "Album © line",

    # Fields intentionally ignored (no Gallo equivalent):
    # album description, gender, publishing rights organisation,
    # rights territories, country, file name, rbt start time
}

# Fields that should be copied to a second Gallo column simultaneously
DUAL_MAP = {
    "genre": "Track genre",          # genre → Album genre (primary) + Track genre (secondary)
    "language*": "Audio language",   # language → Language (primary) + Audio language (secondary)
    "language": "Audio language",
}


def normalise(header):
    """Lowercase + strip for fuzzy key lookup."""
    return str(header).lower().strip() if header else ""


def format_value(value):
    """Convert openpyxl cell values to clean strings/values for writing."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, dt_time):
        # Duration: return as HH:MM:SS string
        return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"
    return value


def load_headers(ws):
    """Return list of header values from row 1."""
    return [cell.value for cell in ws[1]]


def merge(intake_path, gallo_path, intake_sheet=None, gallo_sheet="Metadata",
          output_path=None, dry_run=False):

    if not os.path.exists(intake_path):
        sys.exit(f"ERROR: Intake file not found: {intake_path}")
    if not os.path.exists(gallo_path):
        sys.exit(f"ERROR: Gallo file not found: {gallo_path}")

    wb_intake = openpyxl.load_workbook(intake_path)
    wb_gallo  = openpyxl.load_workbook(gallo_path)

    # Resolve sheet names
    if intake_sheet is None:
        intake_sheet = wb_intake.sheetnames[0]
    if intake_sheet not in wb_intake.sheetnames:
        sys.exit(f"ERROR: Sheet '{intake_sheet}' not found in {intake_path}. "
                 f"Available: {wb_intake.sheetnames}")
    if gallo_sheet not in wb_gallo.sheetnames:
        sys.exit(f"ERROR: Sheet '{gallo_sheet}' not found in {gallo_path}. "
                 f"Available: {wb_gallo.sheetnames}")

    ws_intake = wb_intake[intake_sheet]
    ws_gallo  = wb_gallo[gallo_sheet]

    intake_headers = load_headers(ws_intake)
    gallo_headers  = load_headers(ws_gallo)

    # Build Gallo column index: header → 1-based column number
    gallo_col_index = {h: i+1 for i, h in enumerate(gallo_headers) if h}

    # Build intake → gallo column mapping for this file
    # intake_col_num (1-based) → list of gallo column numbers
    col_mapping = {}
    unmapped = []

    for i, ih in enumerate(intake_headers):
        key = normalise(ih)
        primary = COLUMN_MAP.get(key)
        if primary and primary in gallo_col_index:
            targets = [gallo_col_index[primary]]
            # Add dual-map target if present
            secondary_header = DUAL_MAP.get(key)
            if secondary_header and secondary_header in gallo_col_index:
                targets.append(gallo_col_index[secondary_header])
            col_mapping[i+1] = targets
        else:
            unmapped.append(ih)

    print(f"\n{'DRY RUN — ' if dry_run else ''}Merging: {intake_path}")
    print(f"  Intake sheet : {intake_sheet}")
    print(f"  Gallo sheet  : {gallo_sheet}")
    print(f"  Mapped columns   : {len(col_mapping)}")
    if unmapped:
        print(f"  Ignored columns  : {unmapped}")

    # Iterate intake data rows (skip header)
    rows_added = 0
    for intake_row in ws_intake.iter_rows(min_row=2, values_only=True):
        # Skip entirely empty rows
        if all(v is None for v in intake_row):
            continue

        # Build a dict of gallo_col_number → value
        new_row_data = {}
        for intake_col_num, gallo_col_nums in col_mapping.items():
            raw_val = intake_row[intake_col_num - 1]
            val = format_value(raw_val)
            for gcn in gallo_col_nums:
                new_row_data[gcn] = val

        if dry_run:
            readable = {gallo_headers[gcn-1]: v for gcn, v in new_row_data.items()}
            print(f"\n  Row {rows_added+1}: {readable}")
        else:
            # Append row at the end of the Gallo sheet
            next_row = ws_gallo.max_row + 1
            for gcn, val in new_row_data.items():
                ws_gallo.cell(row=next_row, column=gcn, value=val)

        rows_added += 1

    if not dry_run:
        out = output_path or gallo_path
        wb_gallo.save(out)
        print(f"  Rows appended    : {rows_added}")
        print(f"  Saved to         : {out}")
    else:
        print(f"\n  Would append {rows_added} rows.")

    return rows_added


def main():
    parser = argparse.ArgumentParser(description="CCA Script — merge intake template into Gallo_Metadata_Extract")
    parser.add_argument("--intake",       required=True,  help="Path to intake .xlsx file (e.g. Time Tested.xlsx)")
    parser.add_argument("--gallo",        required=True,  help="Path to Gallo_Metadata_Extract.xlsx")
    parser.add_argument("--intake-sheet", default=None,   help="Sheet name in intake file (default: first sheet)")
    parser.add_argument("--gallo-sheet",  default="Metadata", help="Sheet name in Gallo file (default: Metadata)")
    parser.add_argument("--output",       default=None,   help="Output file path (default: overwrites Gallo file)")
    parser.add_argument("--dry-run",      action="store_true", help="Preview mapped rows without writing")
    args = parser.parse_args()

    merge(
        intake_path=args.intake,
        gallo_path=args.gallo,
        intake_sheet=args.intake_sheet,
        gallo_sheet=args.gallo_sheet,
        output_path=args.output,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
